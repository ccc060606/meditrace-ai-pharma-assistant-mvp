"""Excel exporter — multi-sheet monthly report export."""
import datetime
import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from src.config import Config

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export monthly report to Excel with multiple sheets."""

    def __init__(self, output_dir: str = None):
        self.output_dir = Path(output_dir or Config().export_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export(self, stats: dict, summary: dict, reports: list,
               month_str: str, session=None) -> str:
        """Generate Excel file. Returns file path."""
        wb = Workbook()

        # Sheet styles
        header_font = Font(name="Microsoft YaHei", bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        def style_header(ws, row, cols):
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        def auto_width(ws, min_width=10, max_width=50):
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                max_len = 0
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

        # ── Sheet 1: 月报概览 ────────────────────────────────────
        ws1 = wb.active
        ws1.title = "月报概览"
        overview_data = [
            ("月份", month_str),
            ("拜访总量", stats["total_visits"]),
            ("客户覆盖数", stats["unique_customers"]),
            ("活动数量", stats["activity_count"]),
            ("跟进任务总数", stats["tasks_total"]),
            ("已完成任务", stats["tasks_completed"]),
            ("待处理任务", stats["tasks_pending"]),
            ("完成率", f"{stats['completion_rate']}%"),
            ("", ""),
            ("科室分布", ""),
        ]
        for dept, count in stats.get("department_distribution", {}).items():
            overview_data.append((f"  {dept}", f"{count} 次"))

        overview_data += [
            ("", ""),
            ("高频沟通主题", ""),
        ]
        for t in stats.get("top_topics", [])[:10]:
            overview_data.append((f"  • {t}", ""))

        overview_data += [
            ("", ""),
            ("重点医学问题", ""),
        ]
        for q in stats.get("medical_questions", [])[:10]:
            overview_data.append((f"  • {q}", ""))

        overview_data += [
            ("", ""),
            ("AI 文字总结", ""),
            ("本月工作进展", summary.get("progress_summary", "")),
            ("重点问题", summary.get("key_issues", "")),
            ("未完成事项", summary.get("unfinished_items", "")),
            ("下月计划", summary.get("next_month_plan", "")),
        ]

        for i, (label, value) in enumerate(overview_data, 1):
            ws1.cell(row=i, column=1, value=label).font = Font(name="Microsoft YaHei", bold=True if not label.startswith("  ") else False, size=10)
            ws1.cell(row=i, column=2, value=value).font = Font(name="Microsoft YaHei", size=10)
            if label and not label.startswith("  "):
                for c in [1, 2]:
                    ws1.cell(row=i, column=c).border = thin_border
        ws1.column_dimensions["A"].width = 20
        ws1.column_dimensions["B"].width = 60

        # ── Sheet 2: 日报明细 ────────────────────────────────────
        ws2 = wb.create_sheet("日报明细")
        headers2 = ["ID", "日期", "科室", "客户编号", "沟通主题", "客户反馈",
                     "后续任务", "截止日期", "跟进状态", "医学问题", "活动名称"]
        for col, h in enumerate(headers2, 1):
            ws2.cell(row=1, column=col, value=h)
        style_header(ws2, 1, len(headers2))

        for i, r in enumerate(reports, 2):
            ws2.cell(row=i, column=1, value=r.id)
            ws2.cell(row=i, column=2, value=str(r.date) if r.date else "")
            ws2.cell(row=i, column=3, value=r.department or "")
            ws2.cell(row=i, column=4, value=r.customer_id or "")
            ws2.cell(row=i, column=5, value=r.topic or "")
            ws2.cell(row=i, column=6, value=r.feedback or "")
            ws2.cell(row=i, column=7, value=r.follow_up_task or "")
            ws2.cell(row=i, column=8, value=str(r.task_deadline) if r.task_deadline else "")
            ws2.cell(row=i, column=9, value=r.follow_up_status or "")
            ws2.cell(row=i, column=10, value=r.medical_question or "")
            ws2.cell(row=i, column=11, value=r.activity_name or "")
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=i, column=col).border = thin_border
                ws2.cell(row=i, column=col).font = Font(name="Microsoft YaHei", size=9)
        auto_width(ws2)

        # ── Sheet 3: 客户汇总 ────────────────────────────────────
        ws3 = wb.create_sheet("客户汇总")
        headers3 = ["客户编号", "科室", "拜访次数", "最近沟通"]
        for col, h in enumerate(headers3, 1):
            ws3.cell(row=1, column=col, value=h)
        style_header(ws3, 1, len(headers3))

        customer_data = {}
        for r in reports:
            if r.customer_id:
                if r.customer_id not in customer_data:
                    customer_data[r.customer_id] = {"dept": r.department, "count": 0, "latest": r.date}
                customer_data[r.customer_id]["count"] += 1
                if r.date and (not customer_data[r.customer_id]["latest"] or r.date > customer_data[r.customer_id]["latest"]):
                    customer_data[r.customer_id]["latest"] = r.date

        for i, (cid, cinfo) in enumerate(sorted(customer_data.items()), 2):
            ws3.cell(row=i, column=1, value=cid)
            ws3.cell(row=i, column=2, value=cinfo["dept"] or "")
            ws3.cell(row=i, column=3, value=cinfo["count"])
            ws3.cell(row=i, column=4, value=str(cinfo["latest"]) if cinfo["latest"] else "")
            for col in range(1, len(headers3) + 1):
                ws3.cell(row=i, column=col).border = thin_border
                ws3.cell(row=i, column=col).font = Font(name="Microsoft YaHei", size=10)
        auto_width(ws3)

        # ── Sheet 4: 待跟进事项 ──────────────────────────────────
        ws4 = wb.create_sheet("待跟进事项")
        headers4 = ["ID", "客户编号", "后续任务", "截止日期", "状态"]
        for col, h in enumerate(headers4, 1):
            ws4.cell(row=1, column=col, value=h)
        style_header(ws4, 1, len(headers4))

        row = 2
        for r in reports:
            if r.follow_up_task and r.follow_up_status != "completed":
                ws4.cell(row=row, column=1, value=r.id)
                ws4.cell(row=row, column=2, value=r.customer_id or "")
                ws4.cell(row=row, column=3, value=r.follow_up_task)
                ws4.cell(row=row, column=4, value=str(r.task_deadline) if r.task_deadline else "")
                ws4.cell(row=row, column=5, value=r.follow_up_status or "")
                for col in range(1, len(headers4) + 1):
                    ws4.cell(row=row, column=col).border = thin_border
                    ws4.cell(row=row, column=col).font = Font(name="Microsoft YaHei", size=10)
                row += 1
        auto_width(ws4)

        # ── Sheet 5: 文献列表 ────────────────────────────────────
        ws5 = wb.create_sheet("文献列表")
        headers5 = ["PMID", "标题", "作者", "期刊", "年份", "研究类型", "DOI", "来源链接"]
        for col, h in enumerate(headers5, 1):
            ws5.cell(row=1, column=col, value=h)
        style_header(ws5, 1, len(headers5))

        # Get literature from reports' medical questions
        try:
            from src.models.literature import LiteratureArticle
            s = session
            if s is None:
                from src.db import get_session
                s = get_session()
                own_session = True
            else:
                own_session = False
            try:
                articles = s.query(LiteratureArticle).order_by(LiteratureArticle.year.desc()).all()
                for i, a in enumerate(articles, 2):
                    ws5.cell(row=i, column=1, value=a.pmid or "")
                    ws5.cell(row=i, column=2, value=a.title or "")
                    ws5.cell(row=i, column=3, value=a.authors or "")
                    ws5.cell(row=i, column=4, value=a.journal or "")
                    ws5.cell(row=i, column=5, value=a.year or "")
                    ws5.cell(row=i, column=6, value=a.study_type or "")
                    ws5.cell(row=i, column=7, value=a.doi or "")
                    ws5.cell(row=i, column=8, value=a.source_url or "")
                    for col in range(1, len(headers5) + 1):
                        ws5.cell(row=i, column=col).border = thin_border
                        ws5.cell(row=i, column=col).font = Font(name="Microsoft YaHei", size=9)
            finally:
                if own_session:
                    s.close()
        except Exception:
            pass  # Literature sheet may be empty
        auto_width(ws5)

        # Save
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"月报_{month_str}_{ts}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        logger.info(f"Excel exported to {filepath}")
        return str(filepath)
