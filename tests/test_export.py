"""Tests for Excel and Word exporters."""
import pytest
import datetime
from pathlib import Path
import tempfile
import os

from src.exporters.excel_exporter import ExcelExporter
from src.exporters.word_exporter import WordExporter
from openpyxl import load_workbook


@pytest.fixture
def sample_stats():
    return {
        "month": "2026-05",
        "total_visits": 10,
        "unique_customers": 5,
        "department_distribution": {"心内科": 4, "肿瘤科": 3, "内分泌科": 3},
        "activity_count": 3,
        "activities": ["科室会", "一对一拜访", "学术沙龙"],
        "tasks_total": 5,
        "tasks_completed": 3,
        "tasks_pending": 2,
        "completion_rate": 60.0,
        "top_topics": ["产品介绍", "新药进展", "临床试验"],
        "medical_questions": ["药物安全性？", "疗效对比？"],
        "pending_items": [
            {"customer": "C001", "task": "发送资料", "deadline": "2026-05-30"},
        ],
    }


@pytest.fixture
def sample_summary():
    return {
        "progress_summary": "本月完成10次拜访。",
        "key_issues": "重点问题：药物安全性。",
        "unfinished_items": "2项任务待跟进。",
        "next_month_plan": "继续跟进。",
    }


@pytest.fixture
def sample_reports():
    """Create simple mock report objects."""
    class MockReport:
        def __init__(self, **kwargs):
            for k, v in kwargs.items():
                setattr(self, k, v)
    return [
        MockReport(id=1, date=datetime.date(2026, 5, 5), department="心内科",
                    customer_id="C001", topic="产品介绍", feedback="良好",
                    follow_up_task="T1", task_deadline=datetime.date(2026, 5, 12),
                    follow_up_status="completed", medical_question=None, activity_name="科室会"),
        MockReport(id=2, date=datetime.date(2026, 5, 10), department="肿瘤科",
                    customer_id="C002", topic="新药介绍", feedback="关注疗效",
                    follow_up_task="T2", task_deadline=datetime.date(2026, 5, 20),
                    follow_up_status="pending", medical_question="安全性？", activity_name=None),
    ]


class TestExcelExporter:
    def test_export_creates_file(self, sample_stats, sample_summary, sample_reports, test_db):
        with tempfile.TemporaryDirectory() as tmpdir:
            exporter = ExcelExporter(output_dir=tmpdir)
            filepath = exporter.export(sample_stats, sample_summary, sample_reports, "2026-05", session=test_db)
            assert os.path.exists(filepath)
            assert filepath.endswith(".xlsx")

    def test_export_has_required_sheets(self, sample_stats, sample_summary, sample_reports, test_db):
        with tempfile.TemporaryDirectory() as tmpdir:
            exporter = ExcelExporter(output_dir=tmpdir)
            filepath = exporter.export(sample_stats, sample_summary, sample_reports, "2026-05", session=test_db)
            wb = load_workbook(filepath)
            sheets = wb.sheetnames
            # Must contain at minimum these sheets
            assert "月报概览" in sheets
            assert "日报明细" in sheets
            assert "客户汇总" in sheets
            assert "待跟进事项" in sheets
            assert "文献列表" in sheets

    def test_overview_has_key_data(self, sample_stats, sample_summary, sample_reports, test_db):
        with tempfile.TemporaryDirectory() as tmpdir:
            exporter = ExcelExporter(output_dir=tmpdir)
            filepath = exporter.export(sample_stats, sample_summary, sample_reports, "2026-05", session=test_db)
            wb = load_workbook(filepath)
            ws = wb["月报概览"]
            # Check that key values appear somewhere
            all_text = ""
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        all_text += str(cell.value)
            assert "2026-05" in all_text
            assert "10" in all_text  # total_visits

    def test_export_filename_contains_month(self, sample_stats, sample_summary, sample_reports, test_db):
        with tempfile.TemporaryDirectory() as tmpdir:
            exporter = ExcelExporter(output_dir=tmpdir)
            filepath = exporter.export(sample_stats, sample_summary, sample_reports, "2026-05", session=test_db)
            assert "2026-05" in Path(filepath).name


class TestWordExporter:
    def test_export_creates_file(self, sample_stats, sample_summary, sample_reports, test_db):
        with tempfile.TemporaryDirectory() as tmpdir:
            exporter = WordExporter(output_dir=tmpdir)
            filepath = exporter.export(sample_stats, sample_summary, sample_reports, "2026-05", session=test_db)
            assert os.path.exists(filepath)
            assert filepath.endswith(".docx")

    def test_export_filename_contains_month(self, sample_stats, sample_summary, sample_reports, test_db):
        with tempfile.TemporaryDirectory() as tmpdir:
            exporter = WordExporter(output_dir=tmpdir)
            filepath = exporter.export(sample_stats, sample_summary, sample_reports, "2026-05", session=test_db)
            assert "2026-05" in Path(filepath).name
